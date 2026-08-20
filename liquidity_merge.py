"""
Merge Liquidity Rush / %ofMCAP columns (already computed into the Chartink
Excel workbook by chartink_screener.build_excel -> liquidity_rush.py) into
the quant scorer's results — WITHOUT touching either source file.

Used identically by all three pipelines:
  - main.py           (weekly)   -> one Chartink Excel,  Symbol id column
  - main_daily.py      (daily)    -> one Chartink Excel,  File   id column
  - main_combined.py  (combined) -> TWO Chartink Excels (weekly + daily),
                                     Symbol id column

Nothing here mutates the Chartink Excel or the quant results CSV/DataFrame
passed in — build_liquidity_lookup() only reads, merge_liquidity() returns
a copy, and write_merged_excel() always writes a brand-new file.
"""
import os
import datetime

import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

LIQUIDITY_COLS = [
    "LiquidityRush10days", "%ofMCAP10days",
    "LiquidityRush20days", "%ofMCAP20days",
]

# The "All Tickers" summary tab (see chartink_screener.build_excel) has no
# Liquidity Rush columns at all — it's just a deduplicated ticker list.
SKIP_SHEETS = {"All Tickers"}


def build_liquidity_lookup(excel_paths):
    """
    Reads one or more Chartink Excel workbooks and returns:
        {TICKER: {"LiquidityRush10days": ..., "%ofMCAP10days": ...,
                   "LiquidityRush20days": ..., "%ofMCAP20days": ...}}

    A ticker can appear in several screener sheets inside the same
    workbook (and, for the combined pipeline, in both the weekly and
    daily workbooks) — the first sheet that has a value for it wins, and
    since attach_liquidity_columns() computes these numbers once per
    unique ticker before spreading them across every sheet, that's a
    same-workbook no-op anyway. Sheets with no Ticker header or no
    Liquidity Rush columns (the tiny scratch sheets some screeners
    produce) are skipped automatically.

    excel_paths items that are None or don't exist on disk are skipped
    silently, so callers can pass e.g. [weekly_excel, daily_excel]
    without checking each one first.
    """
    lookup = {}
    for path in excel_paths:
        if not path or not os.path.exists(path):
            continue
        wb = load_workbook(path, data_only=True, read_only=True)
        for sheet_name in wb.sheetnames:
            if sheet_name in SKIP_SHEETS:
                continue
            ws = wb[sheet_name]

            header_row_idx, header = None, None
            for row in ws.iter_rows(min_row=1, max_row=4):
                values = [c.value for c in row]
                if values and values[0] == "Ticker":
                    header_row_idx = row[0].row
                    header = values
                    break
            if header is None or not any(col in header for col in LIQUIDITY_COLS):
                continue  # not a per-screener results sheet, or has no liquidity data

            col_idx = {name: i for i, name in enumerate(header) if name}
            ticker_idx = col_idx["Ticker"]

            for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
                if not row or row[ticker_idx] is None:
                    continue
                ticker = str(row[ticker_idx]).strip().upper()
                if not ticker or ticker in lookup:
                    continue
                entry = {col: row[col_idx[col]] for col in LIQUIDITY_COLS if col in col_idx}
                if entry:
                    lookup[ticker] = entry
        wb.close()
    return lookup


def merge_liquidity(df, lookup, id_col="Symbol"):
    """Returns a COPY of df with the 4 Liquidity Rush / %ofMCAP columns
    added, matched on df[id_col] (case-insensitive, whitespace-stripped).
    Does not modify df in place."""
    df = df.copy()
    for col in LIQUIDITY_COLS:
        df[col] = df[id_col].apply(
            lambda t, _c=col: lookup.get(str(t).strip().upper(), {}).get(_c)
        )
    return df


def write_merged_excel(df, output_path, title):
    """Writes df to a brand-new, lightly-styled Excel file. Never opens
    (let alone edits) any existing file — output_path is created fresh."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Combined"

    hdr_fill = PatternFill("solid", fgColor="1F4E78")
    hdr_font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    now_str = datetime.datetime.now().strftime("%d %b %Y  %H:%M")
    n_cols = max(len(df.columns), 1)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    title_cell = ws.cell(row=1, column=1, value=f"  {title}   |   {len(df)} rows   |   {now_str}")
    title_cell.font = Font(name="Arial", bold=True, size=12, color="FFFFFF")
    title_cell.fill = hdr_fill
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 22

    header_row = 3
    for ci, col_name in enumerate(df.columns, start=1):
        c = ws.cell(row=header_row, column=ci, value=col_name)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = Alignment(horizontal="center", vertical="center")

    for ri, (_, row) in enumerate(df.iterrows(), start=header_row + 1):
        for ci, col_name in enumerate(df.columns, start=1):
            val = row[col_name]
            if pd.isna(val):
                val = None
            ws.cell(row=ri, column=ci, value=val)

    for ci, col_name in enumerate(df.columns, start=1):
        lengths = [len(str(col_name))]
        for v in df[col_name].tolist()[:200]:
            if pd.notna(v):
                lengths.append(len(str(v)))
        ws.column_dimensions[get_column_letter(ci)].width = min(max(max(lengths) + 2, 10), 32)

    ws.freeze_panes = f"A{header_row + 1}"
    if len(df):
        ws.auto_filter.ref = f"A{header_row}:{get_column_letter(n_cols)}{header_row + len(df)}"

    wb.save(output_path)
    return output_path


def merged_output_path(label, base_dir):
    """e.g. nse_setups_weekly_with_liquidity_2026-08-19.xlsx"""
    return os.path.join(
        base_dir,
        f"nse_setups_{label}_with_liquidity_{datetime.date.today().strftime('%Y-%m-%d')}.xlsx",
    )
